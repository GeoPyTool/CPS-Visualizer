"""
Convert additional cross-disciplinary public LA-ICP-MS datasets into 2D matrix
CSVs for the CPS-Visualizer sample suite, alongside the bivalve-shell set.

Datasets (all openly licensed):
2. Archaeology — LA-ICP-MS trace elements of medieval ceramics
   Petrik, J. et al. (2025). Ceramic Trade and Transformation in Medieval
   Madayi. Zenodo. https://doi.org/10.5281/zenodo.16995032  (CC-BY 4.0)
   Sheet "LA_matrix": 14 ceramic samples x ~44 trace elements (ug/g).

3. Biomedical — LA-ICP-MS imaging of tissue sections
   Buchholz, R. et al. (2022). A simple preparation protocol for shipping and
   storage of tissue sections for LA-ICP-MS imaging. Metallomics 14, mfac013.
   Zenodo. https://doi.org/10.5281/zenodo.6204296  (CC-BY 4.0)
   File "LA_Data_C1SA1.csv": 200 x 219 tissue-image grid, P-31 / Fe-57 /
   Zn-64 / Zn-66 intensities.
"""
import os
import csv
import numpy as np
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '..', 'DataSample')
SRC = '/tmp/cps_ds'


def save_matrix(folder, tag, mat):
    mat = np.asarray(mat, dtype=float)
    if mat.size == 0:
        return
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f'{tag}.csv')
    np.savetxt(path, np.round(mat, 4), delimiter=',', fmt='%.4f')
    print(f'  -> {os.path.relpath(path, OUT)}  shape={mat.shape}')


def convert_archaeology():
    """14 ceramics x 44 elements -> per-fabric average matrices (8 fabrics)."""
    xlsx = os.path.join(SRC, 'madayi_chem.xlsx')
    if not os.path.exists(xlsx):
        print('  archaeology: xlsx missing, skip'); return
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb['LA_matrix']
    header = [c.value for c in ws[1]]
    id_col = header.index('ID')
    fabric_col = header.index('Fabric')
    elem_cols = [i for i, h in enumerate(header) if i > id_col and h not in ('Fabric', 'Ware')]
    elem_names = [header[i] for i in elem_cols]
    records = {}   # fabric -> list of rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[id_col] is None:
            continue
        vals = []
        ok = True
        for i in elem_cols:
            v = row[i]
            if v is None:
                ok = False
                break
            try:
                vals.append(float(v))
            except (TypeError, ValueError):
                ok = False
                break
        if not ok:
            continue
        fabric = str(row[fabric_col])
        records.setdefault(fabric, []).append(vals)
    folder = os.path.join(OUT, 'Ceramics_Archaeology')
    n_cols = len(elem_cols)
    for fabric, rows in records.items():
        mat = np.array(rows).mean(axis=0).reshape(1, n_cols)   # 1 x elements
        save_matrix(folder, f'CeramicFabric_{fabric}', mat)
    # also save the full sample x element matrix for cross-sample viewing
    all_rows = [v for rows in records.values() for v in rows]
    save_matrix(folder, 'Ceramics_TraceElements', np.array(all_rows))
    # save element names for reference
    with open(os.path.join(folder, 'elements.txt'), 'w') as f:
        f.write(','.join(elem_names))
    print(f'  archaeology: {len(records)} fabrics, {len(all_rows)} samples x {n_cols} elements')


def convert_biomedical():
    """200x219 tissue LA-ICP-MS imaging grid per isotope."""
    data_dir = os.path.join(SRC, 'msi')
    csv_path = None
    for root, _, files in os.walk(data_dir):
        for f in files:
            if f == 'LA_Data_C1SA1.csv':
                csv_path = os.path.join(root, f)
    if not csv_path:
        print('  biomedical: csv missing, skip'); return
    with open(csv_path) as f:
        r = csv.reader(f)
        h1 = next(r); next(r); h3 = next(r)
        rows = []
        for row in r:
            if not row or not row[0].strip():
                continue
            try:
                rows.append([float(x) if x else 0.0 for x in row[1:]])
            except (ValueError, IndexError):
                continue
    data = np.array(rows)          # n_points x n_cols
    isotopes = list(dict.fromkeys(h3[1:]))
    n_iso = len(isotopes)
    folder = os.path.join(OUT, 'Tissue_Biomedical')
    for k, iso in enumerate(isotopes):
        # columns for this isotope: k, k+n_iso, k+2*n_iso, ...
        idxs = range(k, data.shape[1], n_iso)
        arr = data[:, list(idxs)]
        tag = iso.replace(':', '').replace('-', '_')
        save_matrix(folder, f'Tissue_{tag}', arr.T)   # positions x points
    print('  biomedical: tissue imaging grid per isotope')


def main():
    print('== Archaeology ==')
    convert_archaeology()
    print('== Biomedical ==')
    convert_biomedical()


if __name__ == '__main__':
    main()
