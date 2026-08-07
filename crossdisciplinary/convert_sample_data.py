"""
Convert cross-disciplinary public LA-ICP-MS datasets into 2D matrix CSVs
for the CPS-Visualizer sample suite.

Datasets (all openly licensed):
1. Bivalve shell trace-element LA-ICP-MS profiles (environmental / marine biology)
   de Winter, N. (2025). Trace element profiles and environmental data for
   cultured Cerastoderma edule and Ruditapes philippinarium. Zenodo.
   https://doi.org/10.5281/zenodo.18873283
   Files: B325_int.csv (cockle), G668_sub.csv (clam), min60_2022.csv (sediment core)

2. Hair mercury LA-ICP-MS tracks (biomedical / archaeometry)
   King, C. (2021). DrCharlieKing/Hair_Hg: First release raw data. Zenodo.
   https://doi.org/10.5281/zenodo.5156997
   (19th-century goldminer hair, mercury exposure)

Each 1-D growth-line / track series is reshaped into a 2-D "waterfall" matrix
(rows = segments along the growth axis, columns = within-segment window), which
is the natural 2-D image form for CPS-Visualizer's surface-scan viewer.
"""
import csv
import os
import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = HERE


def read_isotopes(path):
    """Read a 2-header LA-ICP-MS export and return column-name -> array."""
    with open(path) as f:
        r = csv.reader(f)
        h1 = next(r)
        next(r)  # units row
        cols = {name: i for i, name in enumerate(h1) if name.strip()}
        data = {name: [] for name in cols}
        for row in r:
            if not row or not row[0].strip():
                continue
            for name, i in cols.items():
                try:
                    data[name].append(float(row[i]))
                except (ValueError, IndexError):
                    data[name].append(np.nan)
    return {k: np.array(v) for k, v in data.items()}


def waterfall(series, rows=180):
    """Reshape a 1-D series into a 2-D rows x cols matrix (edge-truncated)."""
    s = np.nan_to_num(np.asarray(series, dtype=float), nan=0.0)
    n = len(s)
    if n < rows * 2:
        rows = max(2, n // 2)
    seg = n // rows
    return s[:seg * rows].reshape(rows, seg)


ELEMENT = {
    'Ca43': 'Ca', '23Na/43Ca': 'Na', '25Mg/43Ca': 'Mg',
    '55Mn/43Ca': 'Mn', '88Sr/43Ca': 'Sr', '138Ba/43Ca': 'Ba',
    'C13_CPS': 'C13', 'Au197_CPS': 'Au', 'Hg202_CPS': 'Hg',
    'S32_CPS': 'S32',
}


def save_matrix(folder, tag, series, cols=160):
    mat = waterfall(series, rows=cols)
    if mat.size == 0 or np.nanmax(mat) <= 0:
        return
    os.makedirs(folder, exist_ok=True)
    out = os.path.join(folder, f'{tag}.csv')
    np.savetxt(out, np.round(mat, 4), delimiter=',', fmt='%.4f')
    print(f'  -> {os.path.relpath(out, HERE)}  shape={mat.shape}')


def main():
    # ---- 1. Bivalve shell trace elements (environmental / marine biology) ----
    biv = os.path.join(OUT, 'bivalve_shell')
    for fn in ['B325_int.csv', 'G668_sub.csv', 'min60_2022.csv']:
        src = os.path.join('/tmp/cps_ds', fn)
        if not os.path.exists(src):
            continue
        tag = fn.replace('_int.csv', '').replace('_sub.csv', '').replace('.csv', '')
        data = read_isotopes(src)
        for col, arr in data.items():
            el = ELEMENT.get(col)
            if el is None:
                continue
            save_matrix(biv, f'{tag}{el}', arr)
        print(f'  bivalve: {fn} processed')

    # ---- 2. Hair mercury tracks (biomedical / archaeometry) ----
    hair = os.path.join(OUT, 'hair')
    hair_xlsx = None
    for root, _, files in os.walk('/tmp/cps_ds/hair_hg'):
        for f in files:
            if f.endswith('.xlsx'):
                hair_xlsx = os.path.join(root, f)
    if hair_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(hair_xlsx, data_only=True)
        for sn in wb.sheetnames:
            if sn.startswith('Iolite'):
                ws = wb[sn]
                # find isotope columns (CPS)
                header = [c.value for c in ws[1]]
                idx = {h: i for i, h in enumerate(header) if h and 'CPS' in str(h)}
                for col, i in idx.items():
                    series = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        try:
                            series.append(float(row[i]))
                        except (TypeError, ValueError):
                            series.append(np.nan)
                    series = np.array(series)
                    el = col.split('_')[0]
                    save_matrix(hair, f'hair_{el}', series)
        print('  hair: Iolite exports processed')
    else:
        print('  hair: xlsx not found, skipping')


if __name__ == '__main__':
    main()
